from openpyxl import load_workbook
from random import *

wb = load_workbook("sample2.xlsx")
ws = wb.active # 활성화된 sheet

# cell 데이터를 불러오기
for x in range(1, 11) :
    for y in range(1, 11) :
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# cell 갯쇼ㅜ를 모름
for x in range(1, ws.max_row + 1) :
    for y in range(1, ws.max_column + 1) : 
        print(ws.cell(row=x, column=y).value, end=" ")
    print()