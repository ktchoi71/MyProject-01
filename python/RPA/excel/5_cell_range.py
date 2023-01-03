from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

 # 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11) : 
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] # 영어 column 만 가지고 오기
# for cell in col_B : 
#     print(cell.value)

# col_range = ws["B:C"] # 영어, 수학 column 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1]
# for cell in row_title:
#     print(cell.value, end=" ")

# print()

# row_range = ws[2:6]
# for row in row_range:
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row]
# for row in row_range:
#     for cell in row:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# print(tuple(ws.rows))

# print(tuple(ws.columns))

# for row in tuple(ws.rows):
#     print(row[2].value)

# for row in ws.iter_rows():
#     print(row[2].value)

for row in ws.iter_rows():
    for i in row:
        print(i.value, end=" ")
    print()

wb.save("sample3.xlsx")