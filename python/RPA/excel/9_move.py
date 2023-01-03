from openpyxl import load_workbook

wb = load_workbook("sample3.xlsx")
ws = wb.active

# ws.move_range("B1:C11", rows=0, cols=1)
ws.insert_cols(2)
ws["B1"] = "국어"

wb.save("sample_korean1.xlsx")