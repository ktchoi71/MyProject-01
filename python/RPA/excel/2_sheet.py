from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 sheet 생성

new_ws = wb["NewSheet"] 

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("samplesheet.xlsx")

